"""
app.py - Кибернадзор Веб Приложение
Запуск: python app.py
Открыть: http://localhost:5000
"""

import os
import re
import csv
import json
import time
import base64
import threading
import traceback
import hashlib
import requests
import urllib.parse
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, jsonify, request
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

app = Flask(__name__)

# ════════════════════════════════════════════════════════
# КОНФИГ
# ════════════════════════════════════════════════════════
CONFIG_FILE  = "config.json"
DB_FILE      = "uploaded_urls.json"   # база уже загруженных URL

DEFAULT_CONFIG = {
    "screens_dir":    r"C:\Users\Пользователь\Desktop\knadzor_bot\screens",
    "data_dir":       r"C:\Users\Пользователь\Desktop\knadzor_bot\Links",
    "auth_file":      r"C:\Users\Пользователь\Desktop\knadzor_bot\auth.json",
    "virustotal_key": "3b76e4ac3fa029b32dc22a99d9a488cf6c9a2a73d0d1b20a87e80d2f6c19e1e5",
    "gsb_key":        "AIzaSyA_HwrniaHj9xaIVgu0BI_3McznLXIltD0",
    "knadzor_url":    "https://knadzor.kz/#/links",
    "violation":      "Интернет-мошенничество (Фишинговый сайт)",
    "object_type":    "Интернет-ресурс",
    "resource_type":  "Тематические сайты",
    "language":       "Русский",
    "group":          "оранжевый",
    "reason":         "Фишинговый ресурс",
}


def load_cfg():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding="utf-8") as f:
            cfg = json.load(f)
        for k, v in DEFAULT_CONFIG.items():
            cfg.setdefault(k, v)
        return cfg
    return DEFAULT_CONFIG.copy()


def save_cfg(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ════════════════════════════════════════════════════════
# БАЗА ЗАГРУЖЕННЫХ URL
# ════════════════════════════════════════════════════════

def load_db() -> set:
    """Загружает базу уже загруженных URL."""
    if os.path.exists(DB_FILE):
        with open(DB_FILE, encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_db(urls: set):
    """Сохраняет базу URL."""
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(list(urls), f, ensure_ascii=False)


def add_to_db(url: str):
    """Добавляет URL в базу."""
    db = load_db()
    db.add(url.strip())
    save_db(db)


def is_already_uploaded(url: str) -> bool:
    """Проверяет есть ли URL уже в базе."""
    db = load_db()
    return url.strip() in db


def import_from_excel(xlsx_path: str) -> int:
    """Импортирует URL из Excel в базу. Возвращает количество добавленных."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    if "URL" not in headers:
        raise ValueError("Колонка 'URL' не найдена в Excel")

    url_col = headers.index("URL")
    existing = load_db()
    before = len(existing)

    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_col]
        if url and str(url).startswith("http"):
            existing.add(str(url).strip())

    save_db(existing)
    return len(existing) - before


# ════════════════════════════════════════════════════════
# СОСТОЯНИЕ
# ════════════════════════════════════════════════════════
state = {
    "running": False,
    "task": None,
    "total": 0, "done": 0, "ok": 0, "err": 0, "skip": 0,
    "log": [],
    "links": [],
    "csv_path": None,
    "stats": {"links": 0, "checked": 0, "dangerous": 0, "screens": 0, "uploaded": 0, "errors": 0, "db_size": 0},
}
lock = threading.Lock()


def log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    with lock:
        state["log"].append({"ts": ts, "msg": msg, "level": level})
        if len(state["log"]) > 1000:
            state["log"] = state["log"][-1000:]
    print(f"[{ts}] {msg}")


def set_prog(done, total, ok=None, err=None, skip=None):
    with lock:
        state["done"] = done
        state["total"] = total
        if ok   is not None: state["ok"]   = ok
        if err  is not None: state["err"]  = err
        if skip is not None: state["skip"] = skip


# ════════════════════════════════════════════════════════
# FLASK РОУТЫ
# ════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/status")
def api_status():
    cfg = load_cfg()
    screens = len(list(Path(cfg["screens_dir"]).glob("*.png"))) if Path(cfg["screens_dir"]).exists() else 0
    db_size = len(load_db())
    with lock:
        s = dict(state["stats"])
        s["screens"] = screens
        s["db_size"] = db_size
        return jsonify({
            "running":     state["running"],
            "task":        state["task"],
            "total":       state["total"],
            "done":        state["done"],
            "ok":          state["ok"],
            "err":         state["err"],
            "skip":        state["skip"],
            "log":         state["log"][-100:],
            "stats":       s,
            "links_count": len(state["links"]),
        })


@app.route("/api/config", methods=["GET"])
def api_cfg_get():
    return jsonify(load_cfg())


@app.route("/api/config", methods=["POST"])
def api_cfg_set():
    cfg = load_cfg()
    cfg.update(request.json)
    save_cfg(cfg)
    return jsonify({"ok": True})


@app.route("/api/fetch", methods=["POST"])
def api_fetch():
    if state["running"]:
        return jsonify({"ok": False, "error": "Задача уже выполняется"})
    threading.Thread(target=task_fetch, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/upload_txt", methods=["POST"])
def api_upload_txt():
    if "file" not in request.files:
        return jsonify({"ok": False})
    content = request.files["file"].read().decode("utf-8")
    threading.Thread(target=task_load_txt, args=(content,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/import_excel", methods=["POST"])
def api_import_excel():
    """Импортирует уже загруженные URL из Excel в базу."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Файл не найден"})
    f = request.files["file"]
    tmp_path = "/tmp/import_excel.xlsx"
    f.save(tmp_path)
    try:
        added = import_from_excel(tmp_path)
        db_size = len(load_db())
        log(f"📥 Импорт из Excel: добавлено {added} URL. Всего в базе: {db_size}")
        return jsonify({"ok": True, "added": added, "total": db_size})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/db_info")
def api_db_info():
    db = load_db()
    return jsonify({"size": len(db)})


@app.route("/api/check_and_screenshot", methods=["POST"])
def api_check():
    if state["running"]:
        return jsonify({"ok": False, "error": "Задача уже выполняется"})
    data = request.json or {}
    threading.Thread(target=task_check_and_screenshot,
                     args=(int(data.get("from", 1)), int(data.get("to", 9999))),
                     daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/upload_knadzor", methods=["POST"])
def api_upload():
    if state["running"]:
        return jsonify({"ok": False, "error": "Задача уже выполняется"})
    data = request.json or {}
    threading.Thread(target=task_upload,
                     args=(int(data.get("from", 1)), int(data.get("to", 9999))),
                     daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/auto", methods=["POST"])
def api_auto():
    if state["running"]:
        return jsonify({"ok": False, "error": "Задача уже выполняется"})
    threading.Thread(target=task_auto, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/stop", methods=["POST"])
def api_stop():
    with lock:
        state["running"] = False
    log("⏹ Остановлено", "warn")
    return jsonify({"ok": True})


@app.route("/api/clear_log", methods=["POST"])
def api_clear():
    with lock:
        state["log"] = []
    return jsonify({"ok": True})


# ════════════════════════════════════════════════════════
# ШАГ 1 — СБОР ССЫЛОК
# ════════════════════════════════════════════════════════

def task_fetch():
    with lock:
        state["running"] = True
        state["task"] = "fetch"

    log("📥 Собираю ссылки из всех источников...")
    urls = set()

    # OpenPhish
    try:
        r = requests.get("https://openphish.com/feed.txt", timeout=15)
        lines = [u.strip() for u in r.text.splitlines() if u.strip().startswith("http")]
        urls.update(lines)
        log(f"  ✅ OpenPhish: {len(lines)} ссылок")
    except Exception as e:
        log(f"  ❌ OpenPhish: {e}", "error")

    # PhishTank
    try:
        r = requests.get(
            "http://data.phishtank.com/data/online-valid.json",
            timeout=20, headers={"User-Agent": "phishtank/knadzor"}
        )
        data = r.json()
        pt_urls = [item["url"] for item in data if item.get("url")]
        urls.update(pt_urls)
        log(f"  ✅ PhishTank: {len(pt_urls)} ссылок")
    except Exception as e:
        log(f"  ❌ PhishTank: {e}", "error")

    # URLhaus
    try:
        r = requests.get("https://urlhaus.abuse.ch/downloads/text_online/", timeout=15)
        uh_urls = [u.strip() for u in r.text.splitlines()
                   if u.strip().startswith("http") and not u.startswith("#")]
        urls.update(uh_urls)
        log(f"  ✅ URLhaus: {len(uh_urls)} ссылок")
    except Exception as e:
        log(f"  ❌ URLhaus: {e}", "error")

    all_urls = list(urls)
    log(f"📊 Итого уникальных: {len(all_urls)}")

    # Фильтруем уже загруженные
    db = load_db()
    new_urls = [u for u in all_urls if u not in db]
    skipped = len(all_urls) - len(new_urls)
    log(f"🔍 Уже в кибернадзоре: {skipped} → пропускаю")
    log(f"✅ Новых ссылок для загрузки: {len(new_urls)}")

    _save_links(new_urls)

    with lock:
        state["running"] = False
        state["task"] = None


def task_load_txt(content):
    with lock:
        state["running"] = True
        state["task"] = "fetch"

    log("📂 Загружаю файл...")
    urls = []
    for line in content.strip().splitlines():
        line = line.strip()
        m = re.match(r"^\d+\.\s+(.+)$", line)
        if m:
            urls.append(m.group(1).strip())
        elif line.startswith("http"):
            urls.append(line)

    log(f"📊 Найдено в файле: {len(urls)} ссылок")

    # Фильтруем уже загруженные
    db = load_db()
    new_urls = [u for u in urls if u not in db]
    skipped = len(urls) - len(new_urls)
    log(f"🔍 Уже в кибернадзоре: {skipped} → пропускаю")
    log(f"✅ Новых для загрузки: {len(new_urls)}")

    _save_links(new_urls)

    with lock:
        state["running"] = False
        state["task"] = None


def _save_links(urls):
    cfg = load_cfg()
    os.makedirs(cfg["data_dir"], exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(cfg["data_dir"], f"links_{ts}.csv")

    rows = []
    for i, url in enumerate(urls, 1):
        rows.append({
            "title":          f"Site {i}",
            "url":            url,
            "reason":         cfg["reason"],
            "object_type":    cfg["object_type"],
            "resource_type":  cfg["resource_type"],
            "violation":      cfg["violation"],
            "language":       cfg["language"],
            "group":          cfg["group"],
            "screenshot":     os.path.join(cfg["screens_dir"], f"{i}.png"),
            "publications":   "",
            "subscribers":    "",
            "court_decision": "нет",
            "source":         "mixed",
        })

    if not rows:
        log("⚠️ Нет новых ссылок для сохранения")
        with lock:
            state["links"] = []
            state["stats"]["links"] = 0
        return

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys(), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)

    with lock:
        state["links"] = rows
        state["csv_path"] = csv_path
        state["stats"]["links"] = len(rows)

    log(f"💾 CSV сохранён: {csv_path} ({len(rows)} новых ссылок)")


# ════════════════════════════════════════════════════════
# ШАГ 2 — ПРОВЕРКА + СКРИНШОТЫ (VT API — без капчи)
# ════════════════════════════════════════════════════════

def vt_check_api(url: str, api_key: str) -> dict:
    """Проверка через VT API v3."""
    headers = {"x-apikey": api_key}
    url_id = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")

    r = requests.get(f"https://www.virustotal.com/api/v3/urls/{url_id}",
                     headers=headers, timeout=15)

    if r.status_code == 200:
        attrs = r.json().get("data", {}).get("attributes", {})
        stats = attrs.get("last_analysis_stats", {})
        return {
            "malicious":  stats.get("malicious", 0),
            "suspicious": stats.get("suspicious", 0),
            "harmless":   stats.get("harmless", 0),
            "total":      sum(stats.values()),
        }

    if r.status_code == 404:
        # Отправляем на сканирование
        r2 = requests.post("https://www.virustotal.com/api/v3/urls",
                           headers=headers, data={"url": url}, timeout=15)
        return {"malicious": 0, "suspicious": 0, "harmless": 0, "total": 0, "pending": True}

    return {"error": r.status_code}


def gsb_check_api(url: str, api_key: str) -> bool:
    """Проверка через Google Safe Browsing API."""
    try:
        payload = {
            "client": {"clientId": "knadzor", "clientVersion": "1.0"},
            "threatInfo": {
                "threatTypes": ["MALWARE", "SOCIAL_ENGINEERING", "UNWANTED_SOFTWARE"],
                "platformTypes": ["ANY_PLATFORM"],
                "threatEntryTypes": ["URL"],
                "threatEntries": [{"url": url}]
            }
        }
        r = requests.post(
            f"https://safebrowsing.googleapis.com/v4/threatMatches:find?key={api_key}",
            json=payload, timeout=10
        )
        return bool(r.json().get("matches"))
    except Exception:
        return False


def build_report_html(url, num, vt_data, gsb_dangerous):
    """Генерирует HTML отчёт с результатами проверки."""
    malicious  = vt_data.get("malicious", 0)
    suspicious = vt_data.get("suspicious", 0)
    harmless   = vt_data.get("harmless", 0)
    total      = vt_data.get("total", 0)
    threat     = malicious + suspicious

    if threat > 0 or gsb_dangerous:
        verdict = "ВРЕДОНОСНЫЙ"
        v_color = "#ef4444"
        v_bg    = "rgba(239,68,68,0.1)"
        v_icon  = "🚨"
    else:
        verdict = "НЕ ОБНАРУЖЕНО"
        v_color = "#10b981"
        v_bg    = "rgba(16,185,129,0.1)"
        v_icon  = "✅"

    pct = round(threat / total * 100, 1) if total > 0 else 0
    now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    gsb_txt = "⚠️ Опасный" if gsb_dangerous else "✅ Безопасный"

    return f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&family=JetBrains+Mono&display=swap" rel="stylesheet">
<style>
* {{margin:0;padding:0;box-sizing:border-box;}}
body {{font-family:'Inter',sans-serif;background:#07090f;color:#e2e8f0;width:860px;padding:28px;}}
.hdr {{display:flex;justify-content:space-between;align-items:center;margin-bottom:22px;padding-bottom:16px;border-bottom:1px solid rgba(255,255,255,0.08);}}
.logo {{display:flex;align-items:center;gap:12px;}}
.logo-icon {{width:40px;height:40px;background:linear-gradient(135deg,#1d4ed8,#2563eb);border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:20px;}}
.logo-title {{font-size:17px;font-weight:700;color:#fff;}}
.logo-sub {{font-size:10px;color:#475569;letter-spacing:1px;}}
.meta {{text-align:right;font-family:'JetBrains Mono',monospace;font-size:11px;color:#475569;}}
.url-box {{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:12px 16px;margin-bottom:18px;}}
.url-lbl {{font-size:9px;color:#475569;letter-spacing:2px;text-transform:uppercase;margin-bottom:4px;}}
.url-val {{font-family:'JetBrains Mono',monospace;font-size:12px;color:#93c5fd;word-break:break-all;}}
.verdict {{background:{v_bg};border:1px solid {v_color}44;border-radius:10px;padding:16px 20px;margin-bottom:18px;display:flex;justify-content:space-between;align-items:center;}}
.verdict-left {{display:flex;align-items:center;gap:12px;}}
.verdict-icon {{font-size:30px;}}
.verdict-lbl {{font-size:9px;color:#64748b;letter-spacing:2px;text-transform:uppercase;}}
.verdict-txt {{font-size:20px;font-weight:700;color:{v_color};letter-spacing:1px;}}
.verdict-num {{font-size:30px;font-weight:700;color:{v_color};text-align:right;}}
.verdict-sub {{font-size:11px;color:#64748b;}}
.pbar-bg {{background:rgba(255,255,255,0.06);border-radius:4px;height:7px;overflow:hidden;margin-bottom:18px;}}
.pbar-fill {{height:100%;width:{pct}%;background:{v_color};border-radius:4px;}}
.grid {{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:18px;}}
.card {{border-radius:8px;padding:12px;text-align:center;}}
.card-n {{font-size:24px;font-weight:700;}}
.card-l {{font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-top:3px;}}
.c1{{background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.25);}} .c1 .card-n{{color:#f87171;}}
.c2{{background:rgba(249,115,22,0.1);border:1px solid rgba(249,115,22,0.25);}} .c2 .card-n{{color:#fb923c;}}
.c3{{background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.25);}} .c3 .card-n{{color:#34d399;}}
.c4{{background:rgba(100,116,139,0.1);border:1px solid rgba(100,116,139,0.2);}} .c4 .card-n{{color:#94a3b8;}}
.sources {{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px;}}
.src {{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.07);border-radius:8px;padding:12px;}}
.src-name {{font-size:10px;font-weight:700;letter-spacing:1.5px;color:#64748b;text-transform:uppercase;margin-bottom:6px;}}
.src-val {{font-size:13px;font-weight:600;}}
.footer {{display:flex;justify-content:space-between;padding-top:14px;border-top:1px solid rgba(255,255,255,0.06);font-size:11px;color:#334155;font-family:'JetBrains Mono',monospace;}}
</style></head><body>
<div class="hdr">
  <div class="logo">
    <div class="logo-icon">🦠</div>
    <div><div class="logo-title">Отчёт проверки URL</div><div class="logo-sub">КИБЕРНАДЗОР · МВД РК · АВТОМАТИЗАЦИЯ</div></div>
  </div>
  <div class="meta"><div>SITE №{num}</div><div style="margin-top:3px">{now}</div></div>
</div>
<div class="url-box"><div class="url-lbl">Проверяемый URL</div><div class="url-val">{url}</div></div>
<div class="verdict">
  <div class="verdict-left">
    <span class="verdict-icon">{v_icon}</span>
    <div><div class="verdict-lbl">Итоговый вердикт</div><div class="verdict-txt">{verdict}</div></div>
  </div>
  <div><div class="verdict-num">{threat}<span style="font-size:15px;color:#475569"> / {total}</span></div><div class="verdict-sub">антивирусов обнаружили</div></div>
</div>
<div class="pbar-bg"><div class="pbar-fill"></div></div>
<div class="grid">
  <div class="card c1"><div class="card-n">{malicious}</div><div class="card-l">Вредоносных</div></div>
  <div class="card c2"><div class="card-n">{suspicious}</div><div class="card-l">Подозрительных</div></div>
  <div class="card c3"><div class="card-n">{harmless}</div><div class="card-l">Безопасных</div></div>
  <div class="card c4"><div class="card-n">{total - malicious - suspicious - harmless}</div><div class="card-l">Не проверили</div></div>
</div>
<div class="sources">
  <div class="src"><div class="src-name">VirusTotal</div><div class="src-val" style="color:{'#f87171' if threat>0 else '#34d399'}">{f'⚠️ {threat} обнаружений' if threat>0 else '✅ Чисто'}</div></div>
  <div class="src"><div class="src-name">Google Safe Browsing</div><div class="src-val" style="color:{'#f87171' if gsb_dangerous else '#34d399'}">{gsb_txt}</div></div>
</div>
<div class="footer"><span>VirusTotal API v3 + Google Safe Browsing API v4</span><span>{now}</span></div>
</body></html>"""


def task_check_and_screenshot(num_from=1, num_to=9999):
    with lock:
        state["running"] = True
        state["task"] = "check"
        state["ok"] = state["err"] = state["skip"] = 0
        links = list(state["links"])

    cfg = load_cfg()
    os.makedirs(cfg["screens_dir"], exist_ok=True)

    jobs = []
    for i, r in enumerate(links):
        m = re.search(r"\d+", r["title"])
        if m and num_from <= int(m.group()) <= num_to:
            jobs.append((r, int(m.group())))

    total = len(jobs)
    ok = err = skip = dangerous = 0
    log(f"🔍 Проверяю {total} ссылок через VT API + Google Safe Browsing...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 860, "height": 800})
        page = context.new_page()

        for i, (row, num) in enumerate(jobs, 1):
            if not state["running"]:
                break

            url  = row["url"].strip()
            path = Path(cfg["screens_dir"]) / f"{num}.png"

            if path.exists():
                skip += 1
                set_prog(i, total, ok, err, skip)
                continue

            log(f"[{i}/{total}] Site {num}: {url[:65]}")

            try:
                # VT API проверка
                vt_data = {}
                if cfg.get("virustotal_key"):
                    vt_data = vt_check_api(url, cfg["virustotal_key"])
                    threat = vt_data.get("malicious", 0) + vt_data.get("suspicious", 0)
                    log(f"  VT: {threat}/{vt_data.get('total',0)} обнаружений")

                # GSB проверка
                gsb = False
                if cfg.get("gsb_key"):
                    gsb = gsb_check_api(url, cfg["gsb_key"])
                    log(f"  GSB: {'⚠️ опасный' if gsb else '✅ чистый'}")

                # Генерируем HTML отчёт и делаем скриншот
                html = build_report_html(url, num, vt_data, gsb)
                page.set_content(html, wait_until="networkidle")
                page.wait_for_timeout(1000)
                page.screenshot(path=str(path), full_page=True)

                threat = vt_data.get("malicious", 0) + vt_data.get("suspicious", 0)
                if threat > 0 or gsb:
                    dangerous += 1
                    log(f"  🚨 ОПАСНЫЙ → скриншот сохранён: {num}.png")
                else:
                    log(f"  ✅ Чистый → скриншот сохранён: {num}.png")

                ok += 1

            except Exception as e:
                err += 1
                log(f"  ❌ {e}", "error")

            set_prog(i, total, ok, err, skip)
            with lock:
                state["stats"]["dangerous"] = dangerous
                state["stats"]["errors"] = err

            # Пауза VT API (бесплатный = 4 запроса/мин)
            if i < total and cfg.get("virustotal_key"):
                time.sleep(16)

        browser.close()

    log(f"🔍 Готово: Скриншотов={ok} Ошибок={err} Пропущено={skip}")
    with lock:
        state["running"] = False
        state["task"] = None


# ════════════════════════════════════════════════════════
# ШАГ 3 — ЗАГРУЗКА В КИБЕРНАДЗОР
# ════════════════════════════════════════════════════════

def task_upload(num_from=1, num_to=9999):
    with lock:
        state["running"] = True
        state["task"] = "upload"
        state["ok"] = state["err"] = 0
        links = list(state["links"])

    cfg = load_cfg()

    if not os.path.exists(cfg["auth_file"]):
        log(f"❌ Не найден auth.json: {cfg['auth_file']}", "error")
        with lock:
            state["running"] = False
            state["task"] = None
        return

    # Только ссылки со скриншотами
    jobs = []
    for r in links:
        m = re.search(r"\d+", r["title"])
        if not m:
            continue
        n = int(m.group())
        if num_from <= n <= num_to:
            path = Path(cfg["screens_dir"]) / f"{n}.png"
            if path.exists():
                jobs.append(r)

    total = len(jobs)
    ok = err = dup = 0
    log(f"🚀 Загружаю {total} ссылок (только со скриншотами)...")

    # Загружаем базу перед стартом
    db = load_db()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=150)
        context = browser.new_context(storage_state=cfg["auth_file"])
        page    = context.new_page()

        page.goto(cfg["knadzor_url"], wait_until="domcontentloaded")
        page.wait_for_timeout(2000)

        for i, row in enumerate(jobs, 1):
            if not state["running"]:
                break

            url = row.get("url", "").strip()
            log(f"[{i}/{total}] {row['title']}: {url[:60]}")

            # Дополнительная проверка по базе
            if url in db:
                log(f"  ⏭️ Уже в базе — пропускаю")
                dup += 1
                set_prog(i, total, ok, err, dup)
                continue

            try:
                result = _upload_one(page, row, cfg)

                if result == "ok":
                    ok += 1
                    db.add(url)  # добавляем в базу
                    log(f"  ✅ Загружено")
                elif result == "duplicate":
                    dup += 1
                    db.add(url)  # тоже добавляем — уже есть
                    log(f"  ⏭️ Дубликат в системе — пропускаю")
                else:
                    log(f"  ⚠️ Статус: {result}")

            except Exception as e:
                err += 1
                log(f"  ❌ {e}", "error")
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(500)
                except Exception:
                    pass

            set_prog(i, total, ok, err, dup)
            with lock:
                state["stats"]["uploaded"] = ok
                state["stats"]["errors"]   = err

        # Сохраняем обновлённую базу
        save_db(db)
        log(f"💾 База обновлена: {len(db)} URL")

        try:
            context.storage_state(path=cfg["auth_file"])
        except Exception:
            pass
        browser.close()

    log(f"🚀 Готово: OK={ok} Дубликатов={dup} Ошибок={err}")
    with lock:
        state["running"] = False
        state["task"] = None


def _upload_one(page, row, cfg) -> str:
    """Загружает одну запись. Возвращает 'ok', 'duplicate', 'error'."""
    def clean(v):
        try:
            if pd.isna(v): return ""
        except Exception:
            pass
        return str(v).strip()

    page.locator("button", has_text="Добавить ссылку").first.click()
    page.locator("edit-modal .modal-box--small").wait_for(state="visible", timeout=15000)
    page.wait_for_timeout(300)
    modal = page.locator("edit-modal .modal-box--small").first

    modal.locator('my-input[formcontrolname="title"] input').fill(clean(row["title"]))
    modal.locator('my-input[formcontrolname="url"] input').fill(clean(row["url"]))
    modal.locator('my-textarea[formcontrolname="comment"] textarea').fill(clean(row["reason"]))

    court = clean(row.get("court_decision", "")).lower() in ["1", "true", "да"]
    cb = modal.locator('input[formcontrolname="court_decision"]')
    if court and not cb.is_checked():
        cb.check()

    def select(fc, text):
        if not text: return
        sel = modal.locator(f'my-select-field[formcontrolname="{fc}"] select').first
        sel.wait_for(state="visible", timeout=8000)
        sel.select_option(label=text)
        page.wait_for_timeout(100)

    select("resource_type", clean(row["object_type"]))
    select("resource",      clean(row["resource_type"]))
    select("offense",       clean(row["violation"]))
    select("language",      clean(row["language"]))

    group = clean(row.get("group", "")).lower()
    css_map = {"красный": "attribute--cat-red", "оранжевый": "attribute--cat-orange"}
    css = css_map.get(group)
    if css:
        try:
            gs   = modal.locator("group-select").first
            attr = gs.locator(f"div.{css}").first
            pan  = attr.locator("xpath=../../..").first
            btn  = pan.locator("div.panel__layout.shrink button").first
            btn.click()
            page.wait_for_timeout(200)
        except Exception:
            pass

    num  = re.search(r"\d+", clean(row["title"])).group()
    path = Path(cfg["screens_dir"]) / f"{num}.png"
    if not path.exists():
        raise FileNotFoundError(f"Нет скриншота: {path}")

    fi = modal.locator('input.upload-file[type="file"]').first
    fi.wait_for(state="attached", timeout=8000)
    fi.set_input_files(str(path))
    page.wait_for_timeout(800)

    page.wait_for_timeout(300)
    modal.locator("button.button--primary").first.click()

    # Ждём результата и проверяем дубликат
    for _ in range(30):
        page.wait_for_timeout(100)
        try:
            body = page.inner_text("body")
            if "уже внесен" in body.lower() or "уже внесён" in body.lower():
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(300)
                except Exception:
                    pass
                return "duplicate"
        except Exception:
            pass
        try:
            if not modal.is_visible():
                return "ok"
        except Exception:
            return "ok"

    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    except Exception:
        pass
    return "timeout"


def task_auto():
    log("⚡ АВТО РЕЖИМ: сбор → проверка → загрузка")
    task_fetch()
    time.sleep(2)
    task_check_and_screenshot()
    task_upload()
    log("⚡ АВТО РЕЖИМ ЗАВЕРШЁН!")


# ════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 55)
    print("  🛡️  КИБЕРНАДЗОР")
    print("  Открой: http://localhost:5000")
    print("=" * 55)
    app.run(host="0.0.0.0", port=5000, debug=False)


# ════════════════════════════════════════════════════════
# EXCEL ФУНКЦИИ
# ════════════════════════════════════════════════════════

def detect_resource(raw: str) -> dict:
    if not raw or str(raw).strip().lower() in ["нет", "-", "", "удален", "(аккаунт удален)"]:
        return None
    import re as _re
    v = str(raw).lower().strip()
    raw = str(raw).strip()
    url_match = _re.search(r'https?://[^\s\)\]]+', raw)
    found_url = url_match.group(0) if url_match else None

    if "instagram" in v or "insta" in v:
        ig = _re.search(r'instagram\.com/([^?\s/]+)', v)
        url = found_url or (f"https://www.instagram.com/{ig.group(1)}" if ig else "https://www.instagram.com")
        return {"type":"Instagram","url":url,"object_type":"Страница пользователя","resource_type":"Социальные сети","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "tiktok" in v or "tik tok" in v:
        tt = _re.search(r'tiktok\.com/@([^\s/]+)', v)
        url = found_url or (f"https://www.tiktok.com/@{tt.group(1)}" if tt else "https://www.tiktok.com")
        return {"type":"TikTok","url":url,"object_type":"Страница пользователя","resource_type":"Видеохостинг","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "telegram" in v or "t.me" in v or "телеграм" in v:
        tg = _re.search(r't\.me/([^\s/]+)', v)
        url = found_url or (f"https://t.me/{tg.group(1)}" if tg else "https://t.me")
        return {"type":"Telegram","url":url,"object_type":"Страница пользователя","resource_type":"Мессенджеры","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "whatsapp" in v or "watsapp" in v or "ватсап" in v:
        return {"type":"WhatsApp","url":found_url or "https://www.whatsapp.com","object_type":"Страница пользователя","resource_type":"Мессенджеры","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "olx" in v:
        return {"type":"OLX","url":found_url or "https://www.olx.kz","object_type":"Интернет-ресурс","resource_type":"Тематические сайты","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "facebook" in v or "fb.com" in v:
        return {"type":"Facebook","url":found_url or "https://www.facebook.com","object_type":"Страница пользователя","resource_type":"Социальные сети","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "youtube" in v or "youtu.be" in v:
        return {"type":"YouTube","url":found_url or "https://www.youtube.com","object_type":"Видеоматериал","resource_type":"Видеохостинг","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if "avito" in v:
        return {"type":"Avito","url":found_url or "https://www.avito.ru","object_type":"Интернет-ресурс","resource_type":"Тематические сайты","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    if found_url:
        return {"type":"Сайт","url":found_url,"object_type":"Интернет-ресурс","resource_type":"Тематические сайты","violation":"Интернет-мошенничество (Фишинговый сайт)"}
    return None


@app.route("/api/load_excel", methods=["POST"])
def api_load_excel():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Файл не найден"})
    from openpyxl import load_workbook
    f = request.files["file"]
    tmp = "/tmp/excel_upload.xlsx"
    f.save(tmp)
    try:
        wb = load_workbook(tmp, read_only=True)
        ws = wb.active
        COL_URL = 8; COL_FABULA = 4; COL_METHOD = 7; COL_REGION = 0
        records = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            raw_url    = row[COL_URL]    if len(row) > COL_URL    else None
            raw_fabula = row[COL_FABULA] if len(row) > COL_FABULA else ""
            raw_method = row[COL_METHOD] if len(row) > COL_METHOD else ""
            raw_region = row[COL_REGION] if len(row) > COL_REGION else ""
            res = detect_resource(raw_url)
            if not res:
                continue
            reason = f"Мошенничество. Способ: {raw_method}. Регион: {raw_region}."
            if raw_fabula:
                reason += f" {str(raw_fabula)[:200]}"
            records.append({
                "title":         f"Site {len(records)+1}",
                "url":           res["url"],
                "reason":        reason[:500],
                "object_type":   res["object_type"],
                "resource_type": res["resource_type"],
                "violation":     res["violation"],
                "resource_kind": res["type"],
                "screenshot":    "",
                "publications":  "",
                "subscribers":   "",
                "court_decision":"нет",
                "source":        "Оперативная сводка",
            })
        db = load_db()
        new_records = [r for r in records if r["url"] not in db]
        with lock:
            state["links"] = records
            state["stats"]["links"] = len(records)
        log(f"📊 Excel загружен: {len(records)} записей, новых: {len(new_records)}")
        return jsonify({"ok": True, "total": len(records), "new_count": len(new_records)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/excel_process", methods=["POST"])
def api_excel_process():
    if state["running"]:
        return jsonify({"ok": False, "error": "Задача уже выполняется"})
    threading.Thread(target=task_excel_process, daemon=True).start()
    return jsonify({"ok": True})


def task_excel_process():
    with lock:
        state["running"] = True
        state["task"] = "excel"
        state["ok"] = state["err"] = state["skip"] = 0
        records = list(state["links"])

    cfg = load_cfg()
    screens_dir = Path(cfg["screens_dir"]) / "excel"
    screens_dir.mkdir(parents=True, exist_ok=True)

    db = load_db()
    new_records = [r for r in records if r["url"] not in db]
    skipped = len(records) - len(new_records)
    log(f"📊 Excel обработка: {len(new_records)} новых (пропущено {skipped} из базы)")

    total = len(new_records)
    ok = err = skip = 0

    with sync_playwright() as p:
        # ── Скриншоты ──
        log("📸 Делаю скриншоты аккаунтов...")
        browser = p.chromium.launch(headless=False)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"
        )
        page = ctx.new_page()

        for i, rec in enumerate(new_records, 1):
            if not state["running"]:
                break
            num = re.search(r"\d+", rec["title"]).group()
            path = screens_dir / f"{num}.png"
            rec["screenshot"] = str(path)
            set_prog(i, total, ok, err, skip)

            if path.exists():
                skip += 1
                log(f"[{i}/{total}] SKIP {path.name}")
                continue

            log(f"[{i}/{total}] {rec['resource_kind']}: {rec['url'][:65]}")
            try:
                wait_time = 5000 if rec["resource_kind"] in ("Instagram","TikTok","Facebook") else 3000
                page.goto(rec["url"], wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(wait_time)
                page.evaluate("window.scrollTo(0, 0)")
                page.wait_for_timeout(500)
                page.screenshot(path=str(path), full_page=False)
                ok += 1
                log(f"  ✅ {path.name}")
            except Exception as e:
                err += 1
                log(f"  ❌ {e}", "error")
                try:
                    page.screenshot(path=str(path), full_page=False)
                    ok += 1
                except Exception:
                    pass
            time.sleep(1)

        browser.close()

        # ── Загрузка в кибернадзор ──
        if not os.path.exists(cfg["auth_file"]):
            log(f"❌ Не найден auth.json", "error")
            with lock:
                state["running"] = False
                state["task"] = None
            return

        log("🚀 Загружаю в кибернадзор...")
        ok_u = err_u = dup_u = 0

        browser2 = p.chromium.launch(headless=False, slow_mo=150)
        ctx2 = browser2.new_context(storage_state=cfg["auth_file"])
        page2 = ctx2.new_page()
        page2.goto(cfg["knadzor_url"], wait_until="domcontentloaded")
        page2.wait_for_timeout(2000)

        upload_list = [r for r in new_records if Path(r.get("screenshot","")).exists()]
        for i, rec in enumerate(upload_list, 1):
            if not state["running"]:
                break
            url = rec["url"]
            log(f"[{i}/{len(upload_list)}] {rec['resource_kind']}: {url[:60]}")
            if url in db:
                dup_u += 1
                continue
            try:
                result = _upload_one(page2, rec, cfg)
                if result == "ok":
                    ok_u += 1
                    db.add(url)
                    log(f"  ✅ Загружено")
                elif result == "duplicate":
                    dup_u += 1
                    db.add(url)
                    log(f"  ⏭️ Дубликат")
                else:
                    log(f"  ⚠️ {result}", "warn")
            except Exception as e:
                err_u += 1
                log(f"  ❌ {e}", "error")
                try:
                    page2.keyboard.press("Escape")
                    page2.wait_for_timeout(500)
                except Exception:
                    pass
            with lock:
                state["stats"]["uploaded"] = ok_u

        save_db(db)
        try: ctx2.storage_state(path=cfg["auth_file"])
        except: pass
        browser2.close()

    log(f"✅ Excel завершён: скриншоты OK={ok} ERR={err} | загрузка OK={ok_u} DUP={dup_u} ERR={err_u}")
    with lock:
        state["running"] = False
        state["task"] = None