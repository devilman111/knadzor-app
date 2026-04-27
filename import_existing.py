"""
import_existing.py
Читает Excel с уже загруженными ссылками и сохраняет их в базу.
Запусти ОДИН РАЗ чтобы инициализировать базу.

Запуск: python import_existing.py путь_к_файлу.xlsx
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook

DB_FILE = "uploaded_urls.json"

def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        # Путь по умолчанию
        xlsx_path = input("Введи путь к Excel файлу: ").strip().strip('"')

    if not Path(xlsx_path).exists():
        print(f"❌ Файл не найден: {xlsx_path}")
        return

    print(f"📂 Читаю: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Находим колонку URL
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    try:
        url_col = headers.index("URL")
        print(f"✅ Колонка URL найдена: #{url_col + 1}")
    except ValueError:
        print("❌ Колонка 'URL' не найдена. Доступные колонки:")
        print(headers)
        return

    # Загружаем существующую базу
    existing = set()
    if Path(DB_FILE).exists():
        with open(DB_FILE, encoding="utf-8") as f:
            existing = set(json.load(f))
        print(f"📋 В базе уже есть: {len(existing)} URL")

    # Читаем URL из Excel
    new_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_col]
        if url and str(url).startswith("http"):
            new_urls.add(str(url).strip())

    print(f"📊 Найдено в Excel: {len(new_urls)} URL")

    # Объединяем
    all_urls = existing | new_urls
    added = len(all_urls) - len(existing)

    # Сохраняем
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(list(all_urls), f, ensure_ascii=False)

    print(f"✅ Добавлено новых: {added}")
    print(f"💾 Всего в базе: {len(all_urls)} URL")
    print(f"📁 База сохранена: {DB_FILE}")


if __name__ == "__main__":
    main()